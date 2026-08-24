from __future__ import annotations

import argparse
import shutil
import sqlite3
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import numpy as np
import pandas as pd
import yaml


@dataclass
class ProfileMetadata:
    table: str
    key_columns: list[str]
    key_values: tuple[Any, ...]
    season_column: str
    hour_column: str
    value_column: str
    preserve_columns: list[str]
    preserve_values: tuple[Any, ...]
    output_rule: str


@dataclass
class AccuracySummary:
    rmse: pd.Series
    mae: pd.Series
    rmse_duration: pd.Series


@dataclass
class PyomoAcesResult:
    representatives: pd.DataFrame
    cluster_weights: dict[int, float]
    representative_source_days: dict[int, str]
    cluster_assignments: list[int]
    accuracy: AccuracySummary
    forced_day_ids: list[int]
    objective_value: float


def main() -> None:
    parser = argparse.ArgumentParser(
        description=(
            "Reduce an hourly ACES SQLite database using TSAM clustering or "
            "Pyomo-based duration-curve exceedance-error minimization."
        )
    )
    parser.add_argument(
        "config", nargs="?", default="aces_tsam_8760.yaml", help="YAML configuration"
    )
    parser.add_argument(
        "--inspect", action="store_true", help="Print SQLite tables and columns"
    )
    parser.add_argument(
        "--validate-only",
        action="store_true",
        help="Validate configuration and ACES temporal profiles without clustering",
    )
    args = parser.parse_args()

    config_path = Path(args.config).resolve()
    config = load_config(config_path)
    database = resolve_path(
        config_path.parent, config["project"]["input_database"]
    )
    if args.inspect:
        print(inspect_database(database).to_string(index=False))
        return

    profiles, metadata, seasons, hours = load_aces_profiles(
        database, config, config_path.parent
    )
    print(
        f"Loaded {len(seasons)} days x {len(hours)} hours "
        f"with {len(metadata)} ACES profiles."
    )

    selected_columns, column_weights = select_clustering_columns(
        profiles.columns.tolist(), metadata, config["clustering"]["attributes"]
    )
    print(f"Clustering with {len(selected_columns)} selected profiles.")
    engine = str(config["clustering"]["engine"]).lower()
    validate_clustering_settings(
        profiles=profiles,
        selected_columns=selected_columns,
        metadata=metadata,
        clustering=config["clustering"],
        seasons=seasons,
        hours=hours,
        engine=engine,
    )
    if args.validate_only:
        print(
            f"Configuration, ACES temporal profiles, and {engine} settings "
            "are valid."
        )
        return

    extreme_records = build_extreme_audit_records(
        profiles=profiles,
        selected_columns=selected_columns,
        metadata=metadata,
        clustering=config["clustering"],
        seasons=seasons,
        hours=hours,
        engine=engine,
    )
    if engine == "tsam":
        result = run_tsam(
            profiles[selected_columns],
            column_weights,
            config["clustering"],
        )
        all_result = result.clustering.apply(profiles)
        representatives = all_result.cluster_representatives
        cluster_weights = all_result.cluster_weights
        representative_source_days = map_representative_source_days(
            representatives,
            result.clustering.cluster_centers,
            seasons,
        )
        cluster_assignments = list(all_result.clustering.cluster_assignments)
        accuracy = all_result.accuracy
        forced_day_ids = sorted(
            {int(record["day_id"]) for record in extreme_records}
        )
        objective_value = None
    elif engine in {"pyomo", "pyomo_milp", "pyomo_lp_fixed_days"}:
        pyomo_result = run_pyomo(
            profiles=profiles,
            selected_columns=selected_columns,
            column_weights=column_weights,
            metadata=metadata,
            clustering=config["clustering"],
            seasons=seasons,
            hours=hours,
            engine=engine,
            extreme_records=extreme_records,
        )
        representatives = pyomo_result.representatives
        cluster_weights = pyomo_result.cluster_weights
        representative_source_days = pyomo_result.representative_source_days
        cluster_assignments = pyomo_result.cluster_assignments
        accuracy = pyomo_result.accuracy
        forced_day_ids = pyomo_result.forced_day_ids
        objective_value = pyomo_result.objective_value
    else:
        raise ValueError(
            "clustering.engine must be tsam, pyomo, pyomo_milp, or "
            "pyomo_lp_fixed_days."
        )

    output_database = write_reduced_database(
        source_database=database,
        config=config,
        base_directory=config_path.parent,
        representatives=representatives,
        cluster_weights=cluster_weights,
        representative_source_days=representative_source_days,
        metadata=metadata,
        seasons=seasons,
        hours=hours,
    )
    write_audits(
        output_database.parent,
        seasons,
        config,
        representative_source_days,
        cluster_assignments,
        cluster_weights,
        accuracy,
        engine,
        forced_day_ids,
        objective_value,
        extreme_records,
        selected_columns,
        column_weights,
        metadata,
    )
    print(f"Created reduced ACES database: {output_database}")


def validate_clustering_settings(
    profiles: pd.DataFrame,
    selected_columns: list[str],
    metadata: dict[str, ProfileMetadata],
    clustering: dict[str, Any],
    seasons: list[str],
    hours: list[str],
    engine: str,
) -> None:
    allowed_engines = {
        "tsam",
        "pyomo",
        "pyomo_milp",
        "pyomo_lp_fixed_days",
    }
    if engine not in allowed_engines:
        raise ValueError(
            "clustering.engine must be tsam, pyomo, pyomo_milp, or "
            "pyomo_lp_fixed_days."
        )
    n_representative_days = int(clustering["n_representative_days"])
    if n_representative_days < 1 or n_representative_days > len(seasons):
        raise ValueError(
            "n_representative_days must be between 1 and the number of "
            "complete ACES days."
        )
    if engine == "tsam":
        tsam_config = clustering.get("tsam")
        if not isinstance(tsam_config, dict):
            raise ValueError("Missing clustering.tsam configuration.")
        if int(tsam_config.get("period_duration", 24)) != 24:
            raise ValueError(
                "The ACES database writer requires tsam.period_duration: 24."
            )
        cluster_config = tsam_config.get("cluster", {})
        allowed_methods = {
            "hierarchical",
            "kmeans",
            "kmedoids",
            "kmaxoids",
            "averaging",
            "contiguous",
        }
        if cluster_config.get("method") not in allowed_methods:
            raise ValueError(
                "clustering.tsam.cluster.method must be hierarchical, kmeans, "
                "kmedoids, kmaxoids, averaging, or contiguous."
            )
        representation = cluster_config.get("representation", "medoid")
        representation_type = (
            representation.get("type", "medoid")
            if isinstance(representation, dict)
            else representation
        )
        if representation_type != "medoid":
            raise ValueError(
                "The ACES source-day writer requires TSAM representation.type: "
                "medoid. Synthetic representations such as mean or distribution "
                "cannot be labeled as one historical source day."
            )
        extremes = tsam_config.get("extremes", {})
        if extremes.get("enabled", False) and extremes.get(
            "method", "append"
        ) == "replace":
            raise ValueError(
                "TSAM extremes.method: replace is not supported by the ACES "
                "profile-transfer workflow. Use append or new_cluster."
            )
        segmentation = tsam_config.get("segmentation", {})
        if segmentation.get("enabled", False) and int(
            segmentation.get("n_segments", 24)
        ) != 24:
            raise ValueError(
                "The ACES database writer requires 24 TSAM segments (H00-H23)."
            )
        if bool(
            tsam_config.get("aggregate", {}).get(
                "preserve_column_means", False
            )
        ):
            raise ValueError(
                "Use tsam.aggregate.preserve_column_means: false with medoid "
                "output so representative profiles remain exact historical days."
            )
        build_extreme_audit_records(
            profiles=profiles,
            selected_columns=selected_columns,
            metadata=metadata,
            clustering=clustering,
            seasons=seasons,
            hours=hours,
            engine=engine,
        )
        return

    pyomo_config = clustering.get("pyomo")
    if not isinstance(pyomo_config, dict):
        raise ValueError("Missing clustering.pyomo configuration.")
    solution_method = pyomo_config.get("solution_method", "opt")
    if solution_method not in {"opt", "hybrid_random_weighting"}:
        raise ValueError(
            "clustering.pyomo.solution_method must be opt or "
            "hybrid_random_weighting."
        )
    formulation = pyomo_config.get("formulation", "milp")
    if engine == "pyomo_milp":
        formulation = "milp"
    elif engine == "pyomo_lp_fixed_days":
        formulation = "lp_fixed_days"
    if formulation not in {"milp", "lp_fixed_days"}:
        raise ValueError(
            "clustering.pyomo.formulation must be milp or lp_fixed_days."
        )
    if formulation == "lp_fixed_days":
        fixed_day_ids = pyomo_config.get("fixed_day_ids", [])
        if len(set(fixed_day_ids)) != n_representative_days:
            raise ValueError(
                "lp_fixed_days requires exactly n_representative_days unique "
                "fixed_day_ids."
            )
        if bool(pyomo_config.get("use_integer_weights", False)):
            raise ValueError(
                "lp_fixed_days requires use_integer_weights: false."
            )
    if solution_method == "hybrid_random_weighting":
        if formulation != "milp":
            raise ValueError(
                "hybrid_random_weighting requires formulation: milp."
            )
        if int(pyomo_config.get("n_random_iterations", 50)) < 1:
            raise ValueError("n_random_iterations must be at least 1.")
    elif pyomo_config.get("sampled_candidate_pool_size") is not None:
        raise ValueError(
            "sampled_candidate_pool_size is used only by "
            "hybrid_random_weighting; set it to null for solution_method: opt."
        )

    candidate_method = pyomo_config.get("candidate_reduction_method", "none")
    allowed_candidate_methods = {
        "none",
        "kmeans_nearest_day",
        "ward_nearest_day",
        "extreme_plus_kmeans",
        "extreme_plus_ward",
    }
    if candidate_method not in allowed_candidate_methods:
        raise ValueError(
            "Unsupported clustering.pyomo.candidate_reduction_method."
        )
    candidate_count = pyomo_config.get("n_candidate_days")
    if candidate_method != "none":
        if candidate_count is None:
            raise ValueError(
                "n_candidate_days is required when Pyomo candidate reduction "
                "is enabled."
            )
        if not n_representative_days <= int(candidate_count) <= len(seasons):
            raise ValueError(
                "n_candidate_days must be between n_representative_days and "
                "the number of original days."
            )
    feature_mode = pyomo_config.get(
        "candidate_feature_mode", "chronological_daily_profile"
    )
    if feature_mode not in {
        "chronological_daily_profile",
        "daily_duration_curve",
        "hybrid_profile_and_duration",
        "summary_statistics",
    }:
        raise ValueError("Unsupported clustering.pyomo.candidate_feature_mode.")

    extreme_records = build_extreme_audit_records(
        profiles=profiles,
        selected_columns=selected_columns,
        metadata=metadata,
        clustering=clustering,
        seasons=seasons,
        hours=hours,
        engine=engine,
    )
    forced_day_ids = {
        int(record["day_id"]) for record in extreme_records
    }
    if len(forced_day_ids) > n_representative_days:
        raise ValueError(
            f"Pyomo configuration forces {len(forced_day_ids)} unique days, "
            f"which exceeds n_representative_days={n_representative_days}."
        )
    invalid = sorted(
        day_id
        for day_id in forced_day_ids
        if day_id < 1 or day_id > len(seasons)
    )
    if invalid:
        raise ValueError(f"Invalid Pyomo forced day IDs: {invalid}")


def load_config(path: Path) -> dict[str, Any]:
    with path.open(encoding="utf-8") as handle:
        config = yaml.safe_load(handle)
    if not isinstance(config, dict):
        raise ValueError("The YAML configuration must contain a top-level mapping.")
    for section in ("project", "aces", "clustering"):
        if section not in config:
            raise ValueError(f"Missing required YAML section: {section}")
    return config


def load_aces_profiles(
    database: Path, config: dict[str, Any], base_directory: Path
) -> tuple[pd.DataFrame, dict[str, ProfileMetadata], list[str], list[str]]:
    if not database.exists():
        raise FileNotFoundError(f"Input ACES database not found: {database}")

    calendar = config["aces"]["calendar"]
    with sqlite3.connect(f"file:{database.resolve()}?mode=ro", uri=True) as connection:
        seasons = [
            row[0]
            for row in connection.execute(
                f'SELECT "{calendar["season_column"]}" '
                f'FROM "{calendar["season_table"]}"'
            )
        ]
        hours = [
            row[0]
            for row in connection.execute(
                f'SELECT "{calendar["hour_column"]}" '
                f'FROM "{calendar["hour_table"]}"'
            )
        ]
        seasons = sort_seasons(
            seasons, calendar["season_format"], int(calendar["base_year"])
        )
        hours = sorted(hours, key=parse_hour)
        expected_hours = int(calendar["hours_per_day"])
        if len(hours) != expected_hours:
            raise ValueError(
                f"Expected {expected_hours} time_of_day rows, found {len(hours)}."
            )

        index = build_datetime_index(seasons, hours, calendar)
        frames: list[pd.DataFrame] = []
        metadata: dict[str, ProfileMetadata] = {}
        table_names = {
            row[0]
            for row in connection.execute(
                "SELECT name FROM sqlite_master WHERE type = 'table'"
            )
        }
        for table_config in config["aces"]["temporal_tables"]:
            if not table_config.get("enabled", True):
                continue
            table = table_config["table"]
            if table not in table_names:
                if config.get("validation", {}).get(
                    "fail_on_missing_temporal_table", False
                ):
                    raise ValueError(f"Configured temporal table not found: {table}")
                print(f"Skipping missing table: {table}")
                continue
            frame, table_metadata = load_temporal_table(
                connection, table_config, seasons, hours, index
            )
            if frame is None:
                print(f"Skipping empty table: {table}")
                continue
            frames.append(frame)
            metadata.update(table_metadata)

    if not frames:
        raise ValueError("No enabled ACES temporal table contained data.")
    profiles = pd.concat(frames, axis=1)
    profiles = merge_extra_attributes(
        profiles,
        resolve_path(
            base_directory,
            config["project"].get(
                "extra_attributes_directory", "extra_attributes"
            ),
        ),
    )
    if profiles.isna().any().any():
        missing = profiles.columns[profiles.isna().any()].tolist()
        raise ValueError(
            "Incomplete hourly profiles were found. First missing profiles: "
            f"{missing[:10]}"
        )
    return profiles, metadata, seasons, hours


def load_temporal_table(
    connection: sqlite3.Connection,
    table_config: dict[str, Any],
    seasons: list[str],
    hours: list[str],
    index: pd.DatetimeIndex,
) -> tuple[pd.DataFrame | None, dict[str, ProfileMetadata]]:
    table = table_config["table"]
    key_columns = table_config["key_columns"]
    season_column = table_config["season_column"]
    hour_column = table_config["hour_column"]
    value_column = table_config["value_column"]
    preserve_columns = table_config.get("preserve_columns", [])

    row_count = connection.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0]
    if row_count == 0:
        return None, {}

    quoted_keys = ", ".join(f'"{column}"' for column in key_columns)
    key_rows = connection.execute(
        f'SELECT DISTINCT {quoted_keys} FROM "{table}" ORDER BY {quoted_keys}'
    ).fetchall()
    key_to_position = {tuple(row): position for position, row in enumerate(key_rows)}
    season_to_position = {value: position for position, value in enumerate(seasons)}
    hour_to_position = {value: position for position, value in enumerate(hours)}
    season_mapping = build_season_mapping(
        connection, table, season_column, seasons
    )
    matrix = np.full((len(seasons) * len(hours), len(key_rows)), np.nan)

    selected = key_columns + [season_column, hour_column, value_column]
    query = f'SELECT {", ".join(f"""\"{column}\"""" for column in selected)} FROM "{table}"'
    cursor = connection.execute(query)
    while True:
        rows = cursor.fetchmany(100_000)
        if not rows:
            break
        for row in rows:
            key = tuple(row[: len(key_columns)])
            raw_season = row[len(key_columns)]
            season = season_mapping.get(raw_season)
            hour = row[len(key_columns) + 1]
            value = row[len(key_columns) + 2]
            if season is None or hour not in hour_to_position:
                continue
            row_position = (
                season_to_position[season] * len(hours) + hour_to_position[hour]
            )
            matrix[row_position, key_to_position[key]] = value

    columns = [profile_name(table, key) for key in key_rows]
    frame = pd.DataFrame(matrix, index=index, columns=columns)
    metadata: dict[str, ProfileMetadata] = {}
    for key, column in zip(key_rows, columns):
        preserve_values: tuple[Any, ...] = ()
        if preserve_columns:
            where = " AND ".join(f'"{name}" = ?' for name in key_columns)
            preserve_query = (
                f'SELECT {", ".join(f"""\"{name}\"""" for name in preserve_columns)} '
                f'FROM "{table}" WHERE {where} LIMIT 1'
            )
            preserve_row = connection.execute(preserve_query, tuple(key)).fetchone()
            preserve_values = tuple(preserve_row or [None] * len(preserve_columns))
        metadata[column] = ProfileMetadata(
            table=table,
            key_columns=key_columns,
            key_values=tuple(key),
            season_column=season_column,
            hour_column=hour_column,
            value_column=value_column,
            preserve_columns=preserve_columns,
            preserve_values=preserve_values,
            output_rule=table_config.get("output_rule", "representative_profile"),
        )
    return frame, metadata


def build_season_mapping(
    connection: sqlite3.Connection,
    table: str,
    season_column: str,
    canonical_seasons: list[str],
) -> dict[str, str]:
    raw_seasons = {
        row[0]
        for row in connection.execute(
            f'SELECT DISTINCT "{season_column}" FROM "{table}"'
        )
    }
    canonical = set(canonical_seasons)
    mapping: dict[str, str] = {}
    corrected: list[tuple[str, str]] = []
    for raw_season in raw_seasons:
        if raw_season in canonical:
            mapping[raw_season] = raw_season
            continue
        parts = str(raw_season).split("-")
        swapped = "-".join(reversed(parts)) if len(parts) == 2 else ""
        if swapped in canonical:
            mapping[raw_season] = swapped
            corrected.append((raw_season, swapped))

    mapped = list(mapping.values())
    if len(mapping) != len(raw_seasons) or len(set(mapped)) != len(mapped):
        missing = sorted(raw_seasons - set(mapping))
        raise ValueError(
            f'{table} season labels cannot be mapped uniquely to time_season. '
            f"First unmapped labels: {missing[:10]}"
        )
    if corrected:
        print(
            f"Normalized {len(corrected)} swapped day/month labels in {table}; "
            f"example: {corrected[0][0]} -> {corrected[0][1]}"
        )
    return mapping


def merge_extra_attributes(
    profiles: pd.DataFrame, directory: Path
) -> pd.DataFrame:
    if not directory.exists():
        return profiles
    result = profiles.copy()
    for path in sorted(directory.iterdir()):
        if path.suffix.lower() == ".csv":
            extra = pd.read_csv(path)
        elif path.suffix.lower() in {".xlsx", ".xls"}:
            extra = pd.read_excel(path)
        else:
            continue
        if "timestamp" not in extra:
            raise ValueError(f"Extra attribute file needs timestamp: {path}")
        extra["timestamp"] = pd.to_datetime(extra["timestamp"], errors="raise")
        extra = extra.set_index("timestamp")
        duplicate = set(extra.columns).intersection(result.columns)
        if duplicate:
            raise ValueError(f"Duplicate extra attribute columns: {sorted(duplicate)}")
        result = result.join(extra, how="left")
    return result


def select_clustering_columns(
    columns: list[str],
    metadata: dict[str, ProfileMetadata],
    attributes: dict[str, Any],
) -> tuple[list[str], dict[str, float]]:
    include = attributes.get("include", {})
    exclude = attributes.get("exclude", {})
    included_tables = set(include.get("tables", []))
    included_extra = set(include.get("extra_attributes", []))
    excluded_regions = set(str(value) for value in exclude.get("regions", []))
    excluded_techs = set(str(value) for value in exclude.get("technologies", []))
    excluded_demands = set(str(value) for value in exclude.get("demands", []))
    excluded_columns = set(exclude.get("columns", []))
    defaults = attributes.get("table_default_weights", {})
    overrides = attributes.get("column_weight_overrides", {})

    selected: list[str] = []
    weights: dict[str, float] = {}
    for column in columns:
        item = metadata.get(column)
        if item is None:
            if column not in included_extra:
                continue
            selected.append(column)
            weights[column] = float(overrides.get(column, 1.0))
            continue
        if item.table not in included_tables or column in excluded_columns:
            continue
        keyed = {
            name: str(value)
            for name, value in zip(item.key_columns, item.key_values)
        }
        if keyed.get("regions") in excluded_regions:
            continue
        if keyed.get("tech") in excluded_techs:
            continue
        if keyed.get("demand_name") in excluded_demands:
            continue
        selected.append(column)
        weights[column] = float(
            overrides.get(column, defaults.get(item.table, 1.0))
        )
    if not selected:
        raise ValueError("Attribute filters selected zero clustering profiles.")
    return selected, weights


def run_tsam(
    profiles: pd.DataFrame,
    column_weights: dict[str, float],
    clustering: dict[str, Any],
):
    try:
        import tsam
        from tsam import (
            ClusterConfig,
            Distribution,
            ExtremeConfig,
            MinMaxMean,
            SegmentConfig,
        )
    except ImportError as exc:
        raise RuntimeError(
            "TSAM is not installed in this Python environment. "
            "Run: python -m pip install tsam"
        ) from exc

    tsam_config = clustering["tsam"]
    cluster_data = tsam_config["cluster"]
    if cluster_data.get("normalize_column_means", False):
        tolerance = float(
            tsam_config.get("aggregate", {}).get("numerical_tolerance", 1e-13)
        )
        constant_columns = profiles.columns[
            (profiles.max(axis=0) - profiles.min(axis=0)).abs() <= tolerance
        ].tolist()
        if constant_columns:
            profiles = profiles.drop(columns=constant_columns)
            column_weights = {
                column: weight
                for column, weight in column_weights.items()
                if column in profiles.columns
            }
            print(
                f"Excluded {len(constant_columns)} constant profiles from "
                "distance calculations; they remain in the output database."
            )
    cluster = ClusterConfig(
        method=cluster_data["method"],
        representation=build_representation(
            cluster_data.get("representation"), Distribution, MinMaxMean
        ),
        normalize_column_means=bool(
            cluster_data.get("normalize_column_means", False)
        ),
        use_duration_curves=bool(cluster_data.get("use_duration_curves", False)),
        include_period_sums=bool(cluster_data.get("include_period_sums", False)),
        solver=cluster_data.get("solver", "highs"),
    )
    extreme_data = tsam_config.get("extremes", {})
    extremes = None
    if extreme_data.get("enabled", False):
        extreme_columns = {
            key: extreme_data.get(key, [])
            for key in ("max_value", "min_value", "max_period", "min_period")
        }
        missing = sorted(
            {
                column
                for values in extreme_columns.values()
                for column in values
                if column not in profiles.columns
            }
        )
        if missing:
            raise ValueError(
                "Extreme profile names were not found. Check spelling: "
                f"{missing}"
            )
        if any(extreme_columns.values()):
            extremes = ExtremeConfig(
                method=extreme_data.get("method", "append"),
                **extreme_columns,
            )

    segment_data = tsam_config.get("segmentation", {})
    segments = None
    if segment_data.get("enabled", False):
        if int(segment_data["n_segments"]) != 24:
            raise ValueError(
                "ACES database writing currently requires 24 segments (H00-H23)."
            )
        segments = SegmentConfig(
            n_segments=24,
            representation=build_representation(
                segment_data.get("representation"), Distribution, MinMaxMean
            ),
        )

    aggregate_data = tsam_config.get("aggregate", {})
    return tsam.aggregate(
        profiles,
        n_clusters=int(clustering["n_representative_days"]),
        period_duration=tsam_config.get("period_duration", 24),
        temporal_resolution=tsam_config.get("temporal_resolution", "1h"),
        cluster=cluster,
        segments=segments,
        extremes=extremes,
        weights=column_weights,
        preserve_column_means=bool(
            aggregate_data.get("preserve_column_means", True)
        ),
        rescale_exclude_columns=aggregate_data.get(
            "rescale_exclude_columns", []
        ),
        round_decimals=aggregate_data.get("round_decimals"),
        numerical_tolerance=float(
            aggregate_data.get("numerical_tolerance", 1e-13)
        ),
    )


def run_pyomo(
    profiles: pd.DataFrame,
    selected_columns: list[str],
    column_weights: dict[str, float],
    metadata: dict[str, ProfileMetadata],
    clustering: dict[str, Any],
    seasons: list[str],
    hours: list[str],
    engine: str,
    extreme_records: list[dict[str, Any]] | None = None,
) -> PyomoAcesResult:
    try:
        from repday.candidate_reduction import reduce_candidate_days
        from repday.config import AttributeConfig, SolverConfig
        from repday.model import RepresentativeDayOptimizer
        from repday.preprocessing import prepare_daily_structures
        from repday.search import run_full_opt, run_hybrid_random_weighting
    except ImportError as exc:
        raise RuntimeError(
            "Pyomo representative-day dependencies are not installed. "
            "Run: python -m pip install -r requirements.txt"
        ) from exc

    pyomo_config = clustering.get("pyomo", {})
    n_representative_days = int(clustering["n_representative_days"])
    hours_per_day = len(hours)
    if len(profiles) != len(seasons) * hours_per_day:
        raise ValueError("The ACES profile matrix does not contain complete days.")

    normalization = pyomo_config.get("normalization", "minmax")
    attributes = [
        AttributeConfig(
            name=column,
            column=column,
            weight=float(column_weights[column]),
            normalize=normalization != "none",
            normalization=normalization,
        )
        for column in selected_columns
    ]
    if not any(attribute.weight > 0 for attribute in attributes):
        raise ValueError("Pyomo requires at least one positive attribute weight.")

    hourly = profiles[selected_columns].reset_index()
    hourly["day_id"] = np.repeat(
        np.arange(1, len(seasons) + 1), hours_per_day
    )
    hourly["hour_in_day"] = np.tile(
        np.arange(hours_per_day), len(seasons)
    )
    prepared = prepare_daily_structures(
        hourly=hourly,
        attributes=attributes,
        hours_per_day=hours_per_day,
        require_full_days=True,
    )

    if extreme_records is None:
        extreme_records = build_extreme_audit_records(
            profiles=profiles,
            selected_columns=selected_columns,
            metadata=metadata,
            clustering=clustering,
            seasons=seasons,
            hours=hours,
            engine=engine,
        )
    forced_day_ids = sorted(
        {int(record["day_id"]) for record in extreme_records}
    )
    invalid_forced = [
        day_id
        for day_id in forced_day_ids
        if day_id < 1 or day_id > len(seasons)
    ]
    if invalid_forced:
        raise ValueError(
            f"Pyomo forced_day_ids are outside the ACES year: {invalid_forced}"
        )
    if len(forced_day_ids) > n_representative_days:
        raise ValueError(
            f"Pyomo found {len(forced_day_ids)} unique forced extreme days, "
            f"but n_representative_days is {n_representative_days}. Increase "
            "the representative-day count or remove some extreme rules."
        )

    candidate_method = pyomo_config.get(
        "candidate_reduction_method", "none"
    )
    reduction_result = reduce_candidate_days(
        prepared=prepared,
        attributes=attributes,
        method=candidate_method,
        feature_mode=pyomo_config.get(
            "candidate_feature_mode", "chronological_daily_profile"
        ),
        n_candidate_days=pyomo_config.get("n_candidate_days"),
        include_extreme_days=False,
        random_seed=int(pyomo_config.get("random_seed", 42)),
    )
    candidate_day_ids = sorted(
        set(reduction_result.candidate_day_ids) | set(forced_day_ids)
    )

    solver_data = pyomo_config.get("solver", {})
    solver_config = SolverConfig(
        solver_name=solver_data.get("name", "highs"),
        tee=bool(solver_data.get("tee", False)),
        timelimit_seconds=solver_data.get("time_limit_seconds"),
        mipgap=solver_data.get("mip_gap"),
        threads=solver_data.get("threads"),
    )
    formulation = pyomo_config.get("formulation", "milp")
    if engine == "pyomo_milp":
        formulation = "milp"
    elif engine == "pyomo_lp_fixed_days":
        formulation = "lp_fixed_days"

    solution_method = pyomo_config.get("solution_method", "opt")
    use_integer_weights = bool(
        pyomo_config.get("use_integer_weights", False)
    )
    enforce_positive = bool(
        pyomo_config.get("enforce_positive_weight", True)
    )
    minimum_weight = float(pyomo_config.get("minimum_weight", 0.01))
    fixed_day_ids = [
        int(day_id) for day_id in pyomo_config.get("fixed_day_ids", [])
    ]

    if formulation == "lp_fixed_days":
        if solution_method != "opt":
            raise ValueError(
                "pyomo_lp_fixed_days supports solution_method: opt only."
            )
        missing_extremes = sorted(set(forced_day_ids) - set(fixed_day_ids))
        if missing_extremes:
            raise ValueError(
                "Every forced extreme must appear in fixed_day_ids when using "
                f"pyomo_lp_fixed_days. Missing: {missing_extremes}"
            )

    optimizer = RepresentativeDayOptimizer(
        prepared=prepared,
        attributes=attributes,
        n_representative_days=n_representative_days,
        solver_config=solver_config,
        hours_per_day=hours_per_day,
        use_integer_weights=use_integer_weights,
        n_bins=int(pyomo_config.get("n_bins", 100)),
        forced_day_ids=forced_day_ids,
        candidate_day_ids=candidate_day_ids,
        formulation=formulation,
        fixed_day_ids=fixed_day_ids,
        enforce_positive_weight_for_selected_days=enforce_positive,
        min_weight_if_selected=minimum_weight,
    )

    if solution_method == "opt":
        optimization_result = run_full_opt(optimizer)
    elif solution_method == "hybrid_random_weighting":
        if formulation != "milp":
            raise ValueError(
                "hybrid_random_weighting selects fixed random day sets itself; "
                "use formulation: milp."
            )
        optimization_result = run_hybrid_random_weighting(
            prepared=prepared,
            attributes=attributes,
            solver_config=solver_config,
            n_representative_days=n_representative_days,
            n_bins=int(pyomo_config.get("n_bins", 100)),
            forced_day_ids=forced_day_ids,
            candidate_day_ids=candidate_day_ids,
            n_random_iterations=int(
                pyomo_config.get("n_random_iterations", 50)
            ),
            random_seed=int(pyomo_config.get("random_seed", 42)),
            sampled_candidate_pool_size=pyomo_config.get(
                "sampled_candidate_pool_size"
            ),
            use_integer_weights=use_integer_weights,
            enforce_positive_weight_for_selected_days=enforce_positive,
            min_weight_if_selected=minimum_weight,
        )
    else:
        raise ValueError(
            "clustering.pyomo.solution_method must be opt or "
            "hybrid_random_weighting."
        )

    selected_day_ids = sorted(optimization_result.selected_days)
    missing_weights = [
        day_id
        for day_id in selected_day_ids
        if optimization_result.day_weights.get(day_id, 0.0) <= 0
    ]
    if missing_weights:
        raise ValueError(
            "The ACES writer requires every selected Pyomo day to have a "
            f"positive weight. Zero-weight days: {missing_weights}"
        )

    representatives = build_pyomo_representatives(
        profiles, selected_day_ids, hours_per_day
    )
    cluster_weights = {
        cluster: float(optimization_result.day_weights[day_id])
        for cluster, day_id in enumerate(selected_day_ids)
    }
    representative_source_days = {
        cluster: seasons[day_id - 1]
        for cluster, day_id in enumerate(selected_day_ids)
    }
    cluster_assignments = assign_days_to_pyomo_representatives(
        prepared=prepared,
        attributes=attributes,
        selected_day_ids=selected_day_ids,
    )
    accuracy = calculate_pyomo_accuracy(
        profiles=profiles[selected_columns],
        representatives=representatives[selected_columns],
        cluster_assignments=cluster_assignments,
        cluster_weights=cluster_weights,
    )
    print(
        "Duration-curve optimization selected "
        f"{len(selected_day_ids)} days using Pyomo; "
        f"{len(forced_day_ids)} were forced extreme/manual days."
    )
    return PyomoAcesResult(
        representatives=representatives,
        cluster_weights=cluster_weights,
        representative_source_days=representative_source_days,
        cluster_assignments=cluster_assignments,
        accuracy=accuracy,
        forced_day_ids=forced_day_ids,
        objective_value=float(optimization_result.objective_value),
    )


def build_extreme_audit_records(
    profiles: pd.DataFrame,
    selected_columns: list[str],
    metadata: dict[str, ProfileMetadata],
    clustering: dict[str, Any],
    seasons: list[str],
    hours: list[str],
    engine: str,
) -> list[dict[str, Any]]:
    hours_per_day = len(hours)
    is_pyomo = engine.startswith("pyomo")
    if is_pyomo:
        pyomo_config = clustering.get("pyomo", {})
        extreme_config = pyomo_config.get("extremes", {})
        if (
            extreme_config.get("enabled", False)
            and extreme_config.get(
                "selection_mode", "force_within_count"
            )
            != "force_within_count"
        ):
            raise ValueError(
                "Pyomo extremes.selection_mode must be force_within_count."
            )
        selection_method = "force_within_count"
    else:
        pyomo_config = {}
        extreme_config = clustering.get("tsam", {}).get("extremes", {})
        selection_method = extreme_config.get("method", "append")

    rules = {
        key: list(extreme_config.get(key, []))
        for key in ("max_value", "min_value", "max_period", "min_period")
    } if extreme_config.get("enabled", False) else {
        key: [] for key in ("max_value", "min_value", "max_period", "min_period")
    }
    configured_columns = {
        column for columns in rules.values() for column in columns
    }
    missing = sorted(configured_columns - set(profiles.columns))
    if missing:
        raise ValueError(
            "Pyomo extreme profile names were not found. Check spelling: "
            f"{missing}"
        )

    n_days = len(profiles) // hours_per_day
    records: list[dict[str, Any]] = []
    for rule_name, columns in rules.items():
        for column in columns:
            values = profiles[column].to_numpy(dtype=float).reshape(
                n_days, hours_per_day
            )
            if rule_name == "max_value":
                flat_position = int(np.argmax(values))
                day_position = flat_position // hours_per_day
                hour = hours[flat_position % hours_per_day]
                metric_value = float(values.reshape(-1)[flat_position])
            elif rule_name == "min_value":
                flat_position = int(np.argmin(values))
                day_position = flat_position // hours_per_day
                hour = hours[flat_position % hours_per_day]
                metric_value = float(values.reshape(-1)[flat_position])
            elif rule_name == "max_period":
                period_values = values.sum(axis=1)
                day_position = int(np.argmax(period_values))
                hour = ""
                metric_value = float(period_values[day_position])
            else:
                period_values = values.sum(axis=1)
                day_position = int(np.argmin(period_values))
                hour = ""
                metric_value = float(period_values[day_position])
            records.append(
                extreme_record(
                    engine=engine,
                    selection_method=selection_method,
                    rule=rule_name,
                    profile=column,
                    metadata=metadata.get(column),
                    day_position=day_position,
                    seasons=seasons,
                    hour=hour,
                    metric_value=metric_value,
                    profile_in_clustering=column in selected_columns,
                )
            )

    if is_pyomo and extreme_config.get("include_peak_demand_day", False):
        demand_columns = [
            column
            for column in selected_columns
            if metadata.get(column) is not None
            and metadata[column].table == "DemandSpecificDistribution"
        ]
        if demand_columns:
            demand = normalize_for_extreme_detection(
                profiles[demand_columns]
            ).sum(axis=1)
            daily_values = demand.to_numpy().reshape(
                n_days, hours_per_day
            )
            day_position = int(np.argmax(daily_values.max(axis=1)))
            hour_position = int(np.argmax(daily_values[day_position]))
            records.append(
                extreme_record(
                    engine=engine,
                    selection_method=selection_method,
                    rule="include_peak_demand_day",
                    profile="<all selected DSD profiles>",
                    metadata=None,
                    day_position=day_position,
                    seasons=seasons,
                    hour=hours[hour_position],
                    metric_value=float(
                        daily_values[day_position, hour_position]
                    ),
                    profile_in_clustering=True,
                    table="DemandSpecificDistribution",
                    profile_count=len(demand_columns),
                )
            )

    if is_pyomo and extreme_config.get("include_min_wind_day", False):
        wind_columns = [
            column
            for column in selected_columns
            if metadata.get(column) is not None
            and metadata[column].table
            in {"CapacityFactorTech", "CapacityFactorProcess"}
            and "WIND" in column.upper()
        ]
        if wind_columns:
            wind = profiles[wind_columns].mean(axis=1)
            daily_values = wind.to_numpy().reshape(
                n_days, hours_per_day
            )
            daily_means = daily_values.mean(axis=1)
            day_position = int(np.argmin(daily_means))
            records.append(
                extreme_record(
                    engine=engine,
                    selection_method=selection_method,
                    rule="include_min_wind_day",
                    profile="<all selected wind CF profiles>",
                    metadata=None,
                    day_position=day_position,
                    seasons=seasons,
                    hour="",
                    metric_value=float(daily_means[day_position]),
                    profile_in_clustering=True,
                    table="CapacityFactor",
                    profile_count=len(wind_columns),
                )
            )

    ramp_config = extreme_config.get("include_max_daily_ramp_day", False)
    if is_pyomo and ramp_rule_enabled(ramp_config):
        ramp_options = ramp_config if isinstance(ramp_config, dict) else {}
        direction = str(ramp_options.get("direction", "absolute")).lower()
        if direction not in {"absolute", "upward", "downward"}:
            raise ValueError(
                "include_max_daily_ramp_day.direction must be absolute, "
                "upward, or downward."
            )
        demand_columns = [
            column
            for column in selected_columns
            if metadata.get(column) is not None
            and metadata[column].table == "DemandSpecificDistribution"
        ]
        available_regions = sorted(
            {
                profile_key(metadata[column], "regions")
                for column in demand_columns
                if profile_key(metadata[column], "regions") not in {None, ""}
            }
        )
        configured_regions = ramp_options.get("regions", "all")
        if configured_regions == "all":
            regions = available_regions
        elif isinstance(configured_regions, list):
            regions = [str(region) for region in configured_regions]
        else:
            raise ValueError(
                "include_max_daily_ramp_day.regions must be all or a list."
            )
        missing_regions = sorted(set(regions) - set(available_regions))
        if missing_regions:
            raise ValueError(
                "Demand-ramp regions were not found in selected DSD profiles: "
                f"{missing_regions}"
            )
        available_periods = sorted(
            {
                profile_key(metadata[column], "periods")
                for column in demand_columns
                if profile_key(metadata[column], "periods") not in {None, ""}
            },
            key=lambda value: int(value),
        )
        reference_period = resolve_reference_period(
            ramp_options.get("reference_period", "first_model_period"),
            available_periods,
            "demand ramp",
        )
        for region in regions:
            regional_columns = [
                column
                for column in demand_columns
                if str(profile_key(metadata[column], "regions")) == region
                and str(profile_key(metadata[column], "periods"))
                == str(reference_period)
            ]
            if not regional_columns:
                raise ValueError(
                    f"No selected DSD profiles found for demand ramp region "
                    f"{region} and period {reference_period}."
                )
            demand = normalize_for_extreme_detection(
                profiles[regional_columns]
            ).sum(axis=1)
            daily = demand.to_numpy().reshape(n_days, hours_per_day)
            daily_ramps = calculate_ramps(daily, direction)
            day_position = int(np.argmax(daily_ramps.max(axis=1)))
            ramp_position = int(np.argmax(daily_ramps[day_position]))
            records.append(
                extreme_record(
                    engine=engine,
                    selection_method=selection_method,
                    rule="include_max_daily_ramp_day",
                    profile="<selected regional DSD profiles>",
                    metadata=None,
                    day_position=day_position,
                    seasons=seasons,
                    hour=f"{hours[ramp_position]}->{hours[ramp_position + 1]}",
                    metric_value=float(
                        daily_ramps[day_position, ramp_position]
                    ),
                    profile_in_clustering=True,
                    table="DemandSpecificDistribution",
                    profile_count=len(regional_columns),
                    regions=region,
                    periods=reference_period,
                    demand_name="<all selected>",
                )
            )

    if is_pyomo:
        for day_id in pyomo_config.get("forced_day_ids", []):
            day_position = int(day_id) - 1
            if day_position < 0 or day_position >= len(seasons):
                records.append(
                    {
                        "engine": engine,
                        "selection_method": selection_method,
                        "rule": "manual_forced_day_id",
                        "profile": "",
                        "table": "",
                        "regions": "",
                        "periods": "",
                        "demand_name": "",
                        "tech": "",
                        "profile_keys": "",
                        "day_id": int(day_id),
                        "source_day": "",
                        "hour": "",
                        "metric_value": np.nan,
                        "profile_in_clustering": False,
                        "profile_count": 0,
                    }
                )
                continue
            records.append(
                extreme_record(
                    engine=engine,
                    selection_method=selection_method,
                    rule="manual_forced_day_id",
                    profile="",
                    metadata=None,
                    day_position=day_position,
                    seasons=seasons,
                    hour="",
                    metric_value=np.nan,
                    profile_in_clustering=False,
                )
            )
    return records


def extreme_record(
    engine: str,
    selection_method: str,
    rule: str,
    profile: str,
    metadata: ProfileMetadata | None,
    day_position: int,
    seasons: list[str],
    hour: str,
    metric_value: float,
    profile_in_clustering: bool,
    table: str | None = None,
    profile_count: int = 1,
    regions: Any | None = None,
    periods: Any | None = None,
    demand_name: Any | None = None,
    tech: Any | None = None,
) -> dict[str, Any]:
    keyed = (
        {
            name: value
            for name, value in zip(
                metadata.key_columns, metadata.key_values
            )
        }
        if metadata is not None
        else {}
    )
    return {
        "engine": engine,
        "selection_method": selection_method,
        "rule": rule,
        "profile": profile,
        "table": table or (metadata.table if metadata is not None else ""),
        "regions": keyed.get("regions", "") if regions is None else regions,
        "periods": keyed.get("periods", "") if periods is None else periods,
        "demand_name": (
            keyed.get("demand_name", "")
            if demand_name is None
            else demand_name
        ),
        "tech": keyed.get("tech", "") if tech is None else tech,
        "profile_keys": " | ".join(
            f"{name}={value}" for name, value in keyed.items()
        ),
        "day_id": day_position + 1,
        "source_day": seasons[day_position],
        "hour": hour,
        "metric_value": metric_value,
        "profile_in_clustering": profile_in_clustering,
        "profile_count": profile_count,
    }


def normalize_for_extreme_detection(frame: pd.DataFrame) -> pd.DataFrame:
    minimum = frame.min(axis=0)
    ranges = frame.max(axis=0) - minimum
    ranges = ranges.mask(ranges.abs() <= 1e-13, 1.0)
    return (frame - minimum) / ranges


def profile_key(metadata: ProfileMetadata, name: str) -> Any | None:
    keyed = dict(zip(metadata.key_columns, metadata.key_values))
    return keyed.get(name)


def ramp_rule_enabled(config: Any) -> bool:
    if isinstance(config, dict):
        return bool(config.get("enabled", False))
    return bool(config)


def resolve_reference_period(
    configured_period: Any,
    available_periods: list[Any],
    rule_name: str,
) -> Any:
    if not available_periods:
        raise ValueError(f"No periods are available for {rule_name} selection.")
    if configured_period == "first_model_period":
        return min(available_periods, key=lambda value: int(value))
    matches = [
        period
        for period in available_periods
        if str(period) == str(configured_period)
    ]
    if not matches:
        raise ValueError(
            f"Configured {rule_name} reference period {configured_period} was "
            f"not found. Available periods: {available_periods}"
        )
    return matches[0]


def calculate_ramps(daily_values: np.ndarray, direction: str) -> np.ndarray:
    changes = np.diff(daily_values, axis=1)
    if direction == "absolute":
        return np.abs(changes)
    if direction == "upward":
        return changes
    return -changes


def build_pyomo_representatives(
    profiles: pd.DataFrame,
    selected_day_ids: list[int],
    hours_per_day: int,
) -> pd.DataFrame:
    frames = []
    for cluster, day_id in enumerate(selected_day_ids):
        start = (day_id - 1) * hours_per_day
        frame = profiles.iloc[start : start + hours_per_day].copy()
        frame.index = pd.MultiIndex.from_product(
            [[cluster], range(hours_per_day)],
            names=["cluster", "timestep"],
        )
        frames.append(frame)
    return pd.concat(frames)


def assign_days_to_pyomo_representatives(
    prepared,
    attributes,
    selected_day_ids: list[int],
) -> list[int]:
    feature_parts = []
    for attribute in attributes:
        if attribute.weight <= 0:
            continue
        values = prepared.daily_profiles[attribute.name]
        feature_parts.append(
            values * np.sqrt(float(attribute.weight))
        )
    features = np.concatenate(feature_parts, axis=1)
    day_to_position = {
        day_id: position
        for position, day_id in enumerate(prepared.day_labels)
    }
    representative_features = np.stack(
        [features[day_to_position[day_id]] for day_id in selected_day_ids]
    )
    assignments = []
    for values in features:
        distances = np.sum(
            (representative_features - values.reshape(1, -1)) ** 2,
            axis=1,
        )
        assignments.append(int(np.argmin(distances)))
    for cluster, day_id in enumerate(selected_day_ids):
        assignments[day_to_position[day_id]] = cluster
    return assignments


def calculate_pyomo_accuracy(
    profiles: pd.DataFrame,
    representatives: pd.DataFrame,
    cluster_assignments: list[int],
    cluster_weights: dict[int, float],
) -> AccuracySummary:
    n_days = len(cluster_assignments)
    hours_per_day = len(representatives.loc[0])
    source = profiles.to_numpy(dtype=float).reshape(
        n_days, hours_per_day, len(profiles.columns)
    )
    representative_values = np.stack(
        [
            representatives.loc[cluster].to_numpy(dtype=float)
            for cluster in cluster_weights
        ]
    )
    reconstructed = representative_values[
        np.asarray(cluster_assignments, dtype=int)
    ]
    errors = source - reconstructed
    rmse = np.sqrt(np.mean(errors**2, axis=(0, 1)))
    mae = np.mean(np.abs(errors), axis=(0, 1))

    duration_rmse = []
    target_length = n_days * hours_per_day
    for column_position in range(len(profiles.columns)):
        original = np.sort(source[:, :, column_position].reshape(-1))[::-1]
        values = []
        weights = []
        for cluster, weight in cluster_weights.items():
            values.extend(
                representative_values[cluster, :, column_position].tolist()
            )
            weights.extend([float(weight)] * hours_per_day)
        values_array = np.asarray(values, dtype=float)
        weights_array = np.asarray(weights, dtype=float)
        order = np.argsort(values_array)[::-1]
        sorted_values = values_array[order]
        cumulative = np.cumsum(weights_array[order]) / np.sum(weights_array)
        shares = np.arange(1, target_length + 1) / target_length
        positions = np.searchsorted(cumulative, shares, side="left")
        positions = np.minimum(positions, len(sorted_values) - 1)
        approximated = sorted_values[positions]
        duration_rmse.append(
            float(np.sqrt(np.mean((original - approximated) ** 2)))
        )

    index = pd.Index(profiles.columns, name="attribute")
    return AccuracySummary(
        rmse=pd.Series(rmse, index=index),
        mae=pd.Series(mae, index=index),
        rmse_duration=pd.Series(duration_rmse, index=index),
    )


def build_representation(data, Distribution, MinMaxMean):
    if data is None or isinstance(data, str):
        return data
    representation_type = data.get("type", "mean")
    if representation_type == "distribution":
        return Distribution(
            scope=data.get("scope", "cluster"),
            preserve_minmax=bool(data.get("preserve_minmax", False)),
        )
    if representation_type == "minmax_mean":
        return MinMaxMean(
            max_columns=data.get("max_columns", []),
            min_columns=data.get("min_columns", []),
        )
    return representation_type


def map_representative_source_days(
    representatives: pd.DataFrame,
    cluster_centers: tuple[int, ...] | None,
    seasons: list[str],
) -> dict[int, str]:
    cluster_ids = list(
        dict.fromkeys(representatives.index.get_level_values(0))
    )
    if cluster_centers is None or len(cluster_centers) != len(cluster_ids):
        raise ValueError(
            "TSAM did not return one historical source-day index per "
            "representative cluster. Use representation.type: medoid."
        )
    source_days = {
        cluster: seasons[int(cluster_centers[position])]
        for position, cluster in enumerate(cluster_ids)
    }
    if len(set(source_days.values())) != len(source_days):
        duplicates = sorted(
            {
                day
                for day in source_days.values()
                if list(source_days.values()).count(day) > 1
            }
        )
        raise ValueError(
            "Multiple clusters selected the same historical date, so unique "
            f"ACES season labels cannot be created: {duplicates}"
        )
    return source_days


def write_reduced_database(
    source_database: Path,
    config: dict[str, Any],
    base_directory: Path,
    representatives: pd.DataFrame,
    cluster_weights: dict[int, float],
    representative_source_days: dict[int, str],
    metadata: dict[str, ProfileMetadata],
    seasons: list[str],
    hours: list[str],
) -> Path:
    project = config["project"]
    output = resolve_path(base_directory, project["output_database"])
    output.parent.mkdir(parents=True, exist_ok=True)
    if output.exists():
        if not project.get("overwrite_output", False):
            raise FileExistsError(
                f"Output database already exists: {output}. "
                "Set project.overwrite_output: true to replace it."
            )
        output.unlink()
    template_setting = project.get("output_template_database")
    template_database = (
        resolve_path(base_directory, template_setting)
        if template_setting
        else source_database
    )
    if not template_database.exists():
        raise FileNotFoundError(
            f"Configured output template does not exist: {template_database}"
        )
    shutil.copy2(template_database, output)

    representatives = representatives.copy()
    if not isinstance(representatives.index, pd.MultiIndex):
        raise ValueError("TSAM representatives do not have cluster/timestep index.")
    cluster_ids = list(dict.fromkeys(representatives.index.get_level_values(0)))
    labels = {
        cluster: representative_source_days[cluster] for cluster in cluster_ids
    }
    weight_config = config["aces"]["weights"]
    total_days = len(seasons)

    with sqlite3.connect(output) as connection:
        connection.execute("PRAGMA foreign_keys = OFF")
        calendar = config["aces"]["calendar"]
        connection.execute(f'DELETE FROM "{calendar["season_table"]}"')
        connection.executemany(
            f'INSERT INTO "{calendar["season_table"]}" '
            f'("{calendar["season_column"]}") VALUES (?)',
            [(labels[cluster],) for cluster in cluster_ids],
        )

        connection.execute(f'DELETE FROM "{weight_config["table"]}"')
        segfrac_rows = []
        for cluster in cluster_ids:
            weight = float(cluster_weights[cluster])
            for hour in hours:
                segfrac_rows.append(
                    (
                        labels[cluster],
                        hour,
                        weight / (total_days * len(hours)),
                        f"Representative day weight={weight}",
                    )
                )
        connection.executemany(
            f'INSERT INTO "{weight_config["table"]}" '
            f'("{weight_config["season_column"]}", '
            f'"{weight_config["hour_column"]}", '
            f'"{weight_config["value_column"]}", '
            f'"{weight_config["notes_column"]}") VALUES (?, ?, ?, ?)',
            segfrac_rows,
        )

        if weight_config.get("create_audit_table", True):
            audit_table = weight_config.get(
                "audit_table", "RepresentativeDayWeight"
            )
            connection.execute(f'DROP TABLE IF EXISTS "{audit_table}"')
            connection.execute(
                f'CREATE TABLE "{audit_table}" '
                "(representative_day TEXT PRIMARY KEY, weight REAL NOT NULL)"
            )
            connection.executemany(
                f'INSERT INTO "{audit_table}" VALUES (?, ?)',
                [
                    (labels[cluster], float(cluster_weights[cluster]))
                    for cluster in cluster_ids
                ],
            )

        table_groups: dict[str, list[tuple[str, ProfileMetadata]]] = {}
        for column, item in metadata.items():
            table_groups.setdefault(item.table, []).append((column, item))
        for table, items in table_groups.items():
            connection.execute(f'DELETE FROM "{table}"')
            insert_temporal_profiles(
                connection,
                table,
                items,
                representatives,
                cluster_ids,
                labels,
                cluster_weights,
                hours,
            )
        connection.commit()

    validate_output_database(output, config)
    if project.get("compact_output", True):
        print("Compacting reduced SQLite database...")
        with sqlite3.connect(output) as connection:
            connection.execute("VACUUM")
    return output


def insert_temporal_profiles(
    connection: sqlite3.Connection,
    table: str,
    items: list[tuple[str, ProfileMetadata]],
    representatives: pd.DataFrame,
    cluster_ids: list[int],
    labels: dict[int, str],
    cluster_weights: dict[int, float],
    hours: list[str],
) -> None:
    first = items[0][1]
    insert_columns = (
        first.key_columns
        + [first.season_column, first.hour_column, first.value_column]
        + first.preserve_columns
    )
    placeholders = ", ".join("?" for _ in insert_columns)
    quoted = ", ".join(f'"{column}"' for column in insert_columns)
    query = f'INSERT INTO "{table}" ({quoted}) VALUES ({placeholders})'
    rows = []
    for column, item in items:
        if column not in representatives.columns:
            continue
        values = representatives[column]
        denominator = None
        if item.output_rule == "weighted_and_normalized_to_one":
            denominator = sum(
                float(cluster_weights[cluster])
                * float(values.loc[cluster].sum())
                for cluster in cluster_ids
            )
            if denominator <= 0:
                raise ValueError(f"Cannot normalize nonpositive DSD profile: {column}")
        for cluster in cluster_ids:
            cluster_values = values.loc[cluster].to_numpy()
            for timestep, hour in enumerate(hours):
                value = float(cluster_values[timestep])
                if denominator is not None:
                    value = (
                        value * float(cluster_weights[cluster]) / denominator
                    )
                rows.append(
                    item.key_values
                    + (labels[cluster], hour, value)
                    + item.preserve_values
                )
    connection.executemany(query, rows)


def validate_output_database(database: Path, config: dict[str, Any]) -> None:
    validation = config.get("validation", {})
    weight_config = config["aces"]["weights"]
    with sqlite3.connect(database) as connection:
        if validation.get("require_segfrac_sum_one", True):
            total = connection.execute(
                f'SELECT SUM("{weight_config["value_column"]}") '
                f'FROM "{weight_config["table"]}"'
            ).fetchone()[0]
            tolerance = float(validation.get("segfrac_sum_tolerance", 1e-8))
            if total is None or abs(float(total) - 1.0) > tolerance:
                raise ValueError(f"Output SegFrac sum is {total}, expected 1.")
        if validation.get("require_dsd_sum_one", True):
            rows = connection.execute(
                "SELECT regions, periods, demand_name, SUM(dsd) "
                "FROM DemandSpecificDistribution "
                "GROUP BY regions, periods, demand_name "
                "HAVING ABS(SUM(dsd) - 1.0) > ?",
                (float(validation.get("dsd_sum_tolerance", 1e-8)),),
            ).fetchall()
            if rows:
                raise ValueError(f"Output DSD profiles do not sum to 1: {rows[:5]}")
        integrity = connection.execute("PRAGMA quick_check").fetchone()[0]
        if integrity != "ok":
            raise ValueError(f"Output SQLite quick_check failed: {integrity}")


def write_audits(
    output_directory: Path,
    seasons: list[str],
    config: dict[str, Any],
    representative_source_days: dict[int, str],
    cluster_assignments: list[int],
    cluster_weights: dict[int, float],
    accuracy: AccuracySummary | Any,
    engine: str,
    forced_day_ids: list[int],
    objective_value: float | None,
    extreme_records: list[dict[str, Any]],
    selected_columns: list[str],
    column_weights: dict[str, float],
    metadata: dict[str, ProfileMetadata],
) -> None:
    output_directory.mkdir(parents=True, exist_ok=True)
    assignments = pd.DataFrame(
        {
            "original_day": seasons,
            "cluster": cluster_assignments,
            "representative_day": [
                representative_source_days[cluster]
                for cluster in cluster_assignments
            ],
        }
    )
    assignment_counts = pd.Series(cluster_assignments).value_counts()
    forced_source_days = {
        record["source_day"]
        for record in extreme_records
        if record.get("source_day")
    }
    reason_by_day: dict[str, list[str]] = {}
    for record in extreme_records:
        source_day = record.get("source_day")
        if not source_day:
            continue
        reason = str(record["rule"])
        if record.get("profile"):
            reason = f'{reason}: {record["profile"]}'
        reason_by_day.setdefault(source_day, []).append(reason)

    weights = pd.DataFrame(
        {
            "cluster": list(cluster_weights),
            "source_day": [
                representative_source_days[cluster]
                for cluster in cluster_weights
            ],
            "weight": list(cluster_weights.values()),
            "nearest_assignment_count": [
                int(assignment_counts.get(cluster, 0))
                for cluster in cluster_weights
            ],
        }
    )
    weights["is_forced_extreme"] = weights["source_day"].isin(
        forced_source_days
    )
    weights["forced_by"] = weights["source_day"].map(
        lambda day: " ; ".join(
            dict.fromkeys(reason_by_day.get(day, []))
        )
    )

    extreme_columns = [
        "engine",
        "selection_method",
        "rule",
        "profile",
        "table",
        "regions",
        "periods",
        "demand_name",
        "tech",
        "profile_keys",
        "day_id",
        "source_day",
        "hour",
        "metric_value",
        "profile_in_clustering",
        "profile_count",
    ]
    extremes = pd.DataFrame(extreme_records, columns=extreme_columns)
    source_day_to_cluster = {
        source_day: cluster
        for cluster, source_day in representative_source_days.items()
    }
    if not extremes.empty:
        extremes["selected_as_representative"] = extremes[
            "source_day"
        ].isin(source_day_to_cluster)
        extremes["selected_cluster"] = extremes["source_day"].map(
            source_day_to_cluster
        )
        extremes["representative_weight"] = extremes[
            "selected_cluster"
        ].map(cluster_weights)
    else:
        extremes["selected_as_representative"] = pd.Series(dtype=bool)
        extremes["selected_cluster"] = pd.Series(dtype=float)
        extremes["representative_weight"] = pd.Series(dtype=float)

    profile_rows = []
    for column in selected_columns:
        item = metadata.get(column)
        keyed = (
            {
                name: value
                for name, value in zip(
                    item.key_columns, item.key_values
                )
            }
            if item is not None
            else {}
        )
        profile_rows.append(
            {
                "profile": column,
                "clustering_weight": float(column_weights[column]),
                "table": item.table if item is not None else "extra_attribute",
                "regions": keyed.get("regions", ""),
                "periods": keyed.get("periods", ""),
                "demand_name": keyed.get("demand_name", ""),
                "tech": keyed.get("tech", ""),
                "profile_keys": " | ".join(
                    f"{name}={value}" for name, value in keyed.items()
                ),
            }
        )
    profile_weights = pd.DataFrame(profile_rows)

    accuracy_table = pd.concat(
        {
            "rmse": accuracy.rmse,
            "mae": accuracy.mae,
            "rmse_duration": accuracy.rmse_duration,
        },
        axis=1,
    ).reset_index()
    run_summary = pd.DataFrame(
        [
            {
                "engine": engine,
                "requested_representative_days": config["clustering"][
                    "n_representative_days"
                ],
                "created_representative_days": len(cluster_weights),
                "original_days": len(seasons),
                "total_representative_weight": float(
                    sum(cluster_weights.values())
                ),
                "unique_forced_extreme_days": len(forced_source_days),
                "extreme_reason_rows": len(extremes),
                "pyomo_objective_value": objective_value,
                "pyomo_objective_definition": (
                    "Profile-weighted absolute error between full-year and "
                    "representative-day exceedance shares across duration-curve bins"
                ),
                "important_note": (
                    "Nearest assignment counts are diagnostic. ACES SegFrac "
                    "uses representative weights."
                ),
            }
        ]
    )
    configuration = pd.DataFrame(
        flatten_configuration(config), columns=["parameter", "value"]
    )

    assignments.to_csv(output_directory / "cluster_assignments.csv", index=False)
    weights.to_csv(output_directory / "representative_day_weights.csv", index=False)
    extremes.to_csv(
        output_directory / "extreme_day_provenance.csv", index=False
    )
    if config.get("outputs", {}).get("write_accuracy_metrics", True):
        accuracy_table.to_csv(
            output_directory / f"{engine}_accuracy.csv", index=False
        )
    if engine.startswith("pyomo"):
        forced = pd.DataFrame(
            {
                "day_id": forced_day_ids,
                "source_day": [
                    seasons[day_id - 1] for day_id in forced_day_ids
                ],
            }
        )
        forced.to_csv(output_directory / "pyomo_forced_days.csv", index=False)
        pd.DataFrame(
            [
                {
                    "engine": engine,
                    "objective_value": objective_value,
                    "n_forced_days": len(forced_day_ids),
                }
            ]
        ).to_csv(output_directory / "pyomo_summary.csv", index=False)
    if config.get("outputs", {}).get("write_excel_audit", True):
        excel_path = output_directory / "representative_day_audit.xlsx"
        with pd.ExcelWriter(excel_path, engine="openpyxl") as writer:
            run_summary.to_excel(
                writer, sheet_name="Run_Summary", index=False
            )
            weights.to_excel(
                writer, sheet_name="Representative_Days", index=False
            )
            extremes.to_excel(
                writer, sheet_name="Forced_Extremes", index=False
            )
            assignments.to_excel(
                writer, sheet_name="Day_Assignments", index=False
            )
            profile_weights.to_excel(
                writer, sheet_name="Profile_Weights", index=False
            )
            accuracy_table.to_excel(
                writer, sheet_name="Accuracy", index=False
            )
            configuration.to_excel(
                writer, sheet_name="Configuration", index=False
            )
            format_audit_workbook(writer)


def flatten_configuration(
    value: Any, prefix: str = ""
) -> list[tuple[str, Any]]:
    rows: list[tuple[str, Any]] = []
    if isinstance(value, dict):
        for key, child in value.items():
            path = f"{prefix}.{key}" if prefix else str(key)
            rows.extend(flatten_configuration(child, path))
    elif isinstance(value, list):
        if not value:
            rows.append((prefix, "[]"))
        else:
            for position, child in enumerate(value):
                rows.extend(
                    flatten_configuration(
                        child, f"{prefix}[{position}]"
                    )
                )
    else:
        rows.append((prefix, value))
    return rows


def format_audit_workbook(writer: pd.ExcelWriter) -> None:
    from openpyxl.styles import Font, PatternFill

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for worksheet in writer.book.worksheets:
        worksheet.freeze_panes = "A2"
        if worksheet.max_row > 1 and worksheet.max_column > 0:
            worksheet.auto_filter.ref = worksheet.dimensions
        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        for column_cells in worksheet.columns:
            width = max(
                len(str(cell.value)) if cell.value is not None else 0
                for cell in column_cells
            )
            worksheet.column_dimensions[
                column_cells[0].column_letter
            ].width = min(max(width + 2, 12), 60)


def inspect_database(database: Path) -> pd.DataFrame:
    if not database.exists():
        raise FileNotFoundError(database)
    with sqlite3.connect(f"file:{database.resolve()}?mode=ro", uri=True) as connection:
        tables = pd.read_sql_query(
            "SELECT name FROM sqlite_master WHERE type = 'table' ORDER BY name",
            connection,
        )
        rows = []
        for table in tables["name"]:
            columns = pd.read_sql_query(f'PRAGMA table_info("{table}")', connection)
            rows.extend(
                {
                    "table": table,
                    "column": column["name"],
                    "type": column["type"],
                }
                for _, column in columns.iterrows()
            )
    return pd.DataFrame(rows)


def sort_seasons(values: list[str], date_format: str, year: int) -> list[str]:
    return sorted(
        values,
        key=lambda value: pd.to_datetime(
            f"{year}-{value}", format=f"%Y-{date_format}"
        ),
    )


def parse_hour(value: str) -> int:
    return int(str(value).replace("H", ""))


def build_datetime_index(
    seasons: list[str], hours: list[str], calendar: dict[str, Any]
) -> pd.DatetimeIndex:
    timestamps = []
    year = int(calendar["base_year"])
    date_format = calendar["season_format"]
    for season in seasons:
        day = pd.to_datetime(
            f"{year}-{season}", format=f"%Y-{date_format}"
        )
        timestamps.extend(day + pd.Timedelta(hours=parse_hour(hour)) for hour in hours)
    return pd.DatetimeIndex(timestamps, name="timestamp")


def profile_name(table: str, key_values: tuple[Any, ...]) -> str:
    return "|".join([table, *(str(value) for value in key_values)])


def resolve_path(base: Path, value: str) -> Path:
    path = Path(value)
    return path if path.is_absolute() else base / path


if __name__ == "__main__":
    main()
